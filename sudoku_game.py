import random
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook


class SudokuGame:
    def __init__(self, root):
        root.title("Sudoku Game")
        self.root = root
        self.var2 = BooleanVar()
        self.error_text = [""]
        self.win_text = [""]
        self.lose_text = [""]
        self.dictionary = {}
        self.matrix = []
        self.entries = [[Entry(self.root, bg="#cdd3d4", fg="#333fff", justify=CENTER, width=10)
                         for _ in range(9)]
                        for _ in range(9)]
        self.number = random.randint(1, 6)
        self.btn1 = Button(self.root)
        self.btn2 = Checkbutton(self.root)

    def filling_matrix(self):
        for line in self.entries:
            if isinstance(line, list):
                row = []
                for item in line:
                    row.append(item.get())
                self.matrix.append(row)

    def show_message_about_error(self):
        messagebox.showinfo("Error", self.error_text)

    def show_message_about_win(self):
        messagebox.showinfo("You are win!", self.win_text)

    def show_message_about_lose(self):
        messagebox.showinfo("You are lose", self.lose_text)

    def check_matrix_on_null(self):
        error_counter = 0
        for line in self.matrix:
            for item in line:
                try:
                    if int(item) < 0 or int(item) > 9:
                        error_counter += 1
                        print("Some value exceeds or below the permissible value")
                        self.error_text = ["Error!\nYou are was entered a wrong data"]
                        self.show_message_about_error()
                    else:
                        continue
                except Exception as error:
                    error_counter += 1
                    print(error)
                    self.error_text = ["I have an error!\n It's seems that some cell is null"]
                    self.show_message_about_error()
                    break
            if error_counter == 1:
                break
        return error_counter

    def check_game_on_closed(self):
        if self.root.protocol('WM_DELETE_WINDOW'):
            exit(0)

    def check_matrix_on_rules(self):
        # проверка на правильность судоку по столбцам
        counter = 0
        for j in range(len(self.matrix)):
            for i in range(len(self.matrix)):
                col_elem = self.matrix[i][j]
                for k in range(i + 1, len(self.matrix)):
                    if col_elem == self.matrix[k][j]:
                        counter += 1
                        print(f"Have error in {k + 1} row and {j + 1} column")
                        self.error_text = [f"Have error in {k + 1} row and {j + 1} column"]
                        self.show_message_about_error()
                        break
                if counter != 0:
                    return None

        # проверка на правильность судоку по строкам
        # ( если ошибка в столбцах не зафиксировалась )
        for i in range(len(self.matrix)):
            for j in range(len(self.matrix)):
                row_elem = self.matrix[i][j]
                for k in range(j + 1, len(self.matrix)):
                    if row_elem == self.matrix[i][k]:
                        counter += 1
                        print(f"Have error in {k + 1} row and {i} column")
                        self.error_text = [f"Have error in {k + 1} row and {j + 1} column"]
                        self.show_message_about_error()
                        break
                if counter != 0:
                    return None
        return counter

    def get_matrix(self):
        return self.matrix

    def get_data_from_excel(self):
        wb = load_workbook('sudoku.xlsx')
        sheet = wb[f'Лист{self.number}']
        for i in range(1, 10):
            self.dictionary[i - 1] = []
            for j in range(1, 10):
                self.dictionary[i - 1].append(sheet.cell(row=i, column=j).value)

    def result(self):
        count_of_errors_type_one = self.check_matrix_on_null()
        count_of_errors_type_two = self.check_matrix_on_rules()

        if count_of_errors_type_one == 0 and count_of_errors_type_two == 0:
            self.win_text = ["Congration!\nYou are win"]
            self.show_message_about_win()
        elif count_of_errors_type_two != 0 and count_of_errors_type_one == 0:
            self.lose_text = ["You are lose\nTry again!"]
            self.show_message_about_lose()

    # устанавливаем значения внутри сетки из excel-файла
    def set_grid(self, col, row):
        for i, line in enumerate(self.entries):
            for ind, item in enumerate(line):
                # Устанавливаем положение каждого виджета
                item.grid(row=i, column=ind)
                for key, value in self.dictionary.items():
                    if key == i:
                        for it in value:
                            if ind == value.index(it) and isinstance(it, int):
                                item.config(fg="#000000")
                                item.insert(0, it)

        # При нажатии на кнопку - будут выводиться заполненные пользователем элементы
        # В том числе должны выводиться вписанные заранее элементы
        self.btn1 = Button(self.root, text="Check!", background="#353", foreground="#ccc", padx="10",
                           pady="10", font="13",
                           command=lambda: self.filling_matrix())
        self.btn1.place(x=200, y=200)

        self.btn2 = Checkbutton(self.root, text="Get result",
                                variable=self.var2,
                                onvalue=True, offvalue=False,
                                command=self.result)
        self.btn2.place(x=300, y=200)